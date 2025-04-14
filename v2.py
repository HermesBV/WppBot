import os
import openpyxl
import time
import random
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from openpyxl.styles import PatternFill

# Configuración
ARCHIVO_EXCEL = "contactos.xlsx"
MENSAJE_BASE = """Hola {nombre}!
Te hablo desde el S.O.S por la recorrida para ingresantes en Económicas (Av. Córdoba 2122) a la que te anotaste.
Es mañana *martes a las 18hrs*. El punto de encuentro va a ser en nuestra mesita, entrando por la puerta principal a la izquierda.
Igualmente va a haber estudiantes en la entrada con la remera verde del S.O.S para indicarte cómo llegar.
Cualquier cosa avisame, saludos!"""
MENSAJE_SIN_NOMBRE = """Hola!
Te hablo desde el S.O.S por la recorrida para ingresantes en Económicas (Av. Córdoba 2122) a la que te anotaste.
Es mañana *martes a las 18hrs*. El punto de encuentro va a ser en nuestra mesita, entrando por la puerta principal a la izquierda.
Igualmente va a haber estudiantes en la entrada con la remera verde del S.O.S para indicarte cómo llegar.
Cualquier cosa avisame, saludos!"""
COLOR_ERROR = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Configuración de Chrome
chrome_options = Options()
chrome_options.add_argument("--user-data-dir=C:\\Temp\\ChromeProfile")
chrome_options.add_argument("--start-maximized")
chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])

service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=chrome_options)

def encontrar_ultima_fila(hoja):
    max_fila = hoja.max_row
    while max_fila > 1:
        celda = hoja.cell(row=max_fila, column=1).value
        if celda is not None and str(celda).strip() != "":
            return max_fila
        max_fila -= 1
    return 1

def marcar_error(fila, hoja):
    hoja.cell(row=fila, column=3, value="X").fill = COLOR_ERROR

def marcar_exito(fila, hoja):
    hoja.cell(row=fila, column=3, value="✓")

def validar_numero(numero):
    numero = str(numero).strip().replace(" ", "").replace("-", "")
    if not numero:
        return False
    if len(numero) < 8:  # Número mínimo incluyendo código de país
        return False
    return numero if numero.startswith("+") else f"+{numero}"

def enviar_mensaje(numero, nombre):
    try:
        driver.get(f"https://web.whatsapp.com/send?phone={numero}")
        
        # Esperar carga de la conversación
        input_box = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//div[@role="textbox"][@data-tab="10"]'))
        )
        
        # Construir mensaje
        mensaje = MENSAJE_BASE.format(nombre=nombre.strip()) if nombre.strip() else MENSAJE_SIN_NOMBRE
        
        # Enviar mensaje
        for line in mensaje.split('\n'):
            input_box.send_keys(line)
            input_box.send_keys(Keys.SHIFT + Keys.ENTER)
        input_box.send_keys(Keys.ENTER)
        
        # Verificación de envío
        try:
            WebDriverWait(driver, 12).until(
                EC.presence_of_element_located((By.XPATH, '//span[@data-testid="msg-dblcheck" or @data-testid="msg-time"]'))
            )
            print(f"✓✓ Envío confirmado a {numero}")
            return True
        except:
            # Verificar si el mensaje persiste en el input (envío fallido)
            if mensaje.split('\n')[0] in input_box.text:
                print(f"✗ Error de envío a {numero}")
                return False
            print(f"✓ Envío probablemente exitoso a {numero}")
            return True
            
    except Exception as e:
        print(f"✗ Error crítico con {numero}: {str(e)}")
        return False

try:
    # Esperar carga inicial de WhatsApp Web
    driver.get("https://web.whatsapp.com")
    WebDriverWait(driver, 45).until(
        EC.presence_of_element_located((By.XPATH, '//div[@role="textbox"]'))
    )
    
    # Cargar Excel
    wb = openpyxl.load_workbook(ARCHIVO_EXCEL)
    hoja = wb.active
    
    # Configurar columna Estado
    if hoja.cell(row=1, column=3).value != "Estado":
        hoja.cell(row=1, column=3, value="Estado")
    
    # Determinar rango real de datos
    ultima_fila = encontrar_ultima_fila(hoja)
    print(f"📖 Detected last row with data: {ultima_fila}")
    
    contador = 0
    
    # Procesar solo filas con datos
    for fila_idx in range(2, ultima_fila + 1):
        numero_crudo = hoja.cell(row=fila_idx, column=1).value
        nombre = str(hoja.cell(row=fila_idx, column=2).value).strip() if hoja.cell(row=fila_idx, column=2).value else ""
        
        # Validar número
        numero_validado = validar_numero(numero_crudo)
        if not numero_validado:
            marcar_error(fila_idx, hoja)
            print(f"✗ Número inválido en fila {fila_idx}: {numero_crudo}")
            wb.save(ARCHIVO_EXCEL)
            continue
            
        # Enviar mensaje
        if enviar_mensaje(numero_validado, nombre):
            marcar_exito(fila_idx, hoja)
        else:
            marcar_error(fila_idx, hoja)
        
        wb.save(ARCHIVO_EXCEL)
        contador += 1
        
        # Pausas estratégicas
        if contador % 50 == 0:
            print(f"⏳ Pausa de seguridad de 5 minutos...")
            time.sleep(300)
        else:
            time.sleep(random.uniform(10, 18))

finally:
    try:
        wb.save(ARCHIVO_EXCEL)
        wb.close()
    except Exception as e:
        print(f"⚠ Error guardando Excel: {str(e)}")
    driver.quit()
    print("✅ Proceso finalizado correctamente")