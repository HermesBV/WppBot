import os
import openpyxl
import time
import random
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options

# Configuración
ARCHIVO_CONFIG = "config.xlsx"

# Configuración de Chrome
chrome_options = Options()
chrome_options.add_argument("--user-data-dir=C:\\Temp\\ChromeProfile")
chrome_options.add_argument("--start-maximized")
chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])

service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=chrome_options)

def cargar_tiempos_espera(wb):
    if "Sleeptime" not in wb.sheetnames:
        raise ValueError("No existe la hoja 'Sleeptime'")
    
    hoja_sleeptime = wb["Sleeptime"]
    min_pausa = hoja_sleeptime['E2'].value
    max_pausa = hoja_sleeptime['F2'].value
    pausa_larga = hoja_sleeptime['E3'].value
    
    if None in (min_pausa, max_pausa, pausa_larga):
        raise ValueError("Faltan valores en la hoja 'Sleeptime'")
    
    return float(min_pausa), float(max_pausa), float(pausa_larga)

def encontrar_ultima_fila(hoja):
    max_fila = hoja.max_row
    while max_fila > 1:
        celda = hoja.cell(row=max_fila, column=1).value
        if celda is not None and str(celda).strip() != "":
            return max_fila
        max_fila -= 1
    return 1

def marcar_error(fila, hoja):
    hoja.cell(row=fila, column=3, value="X")

def marcar_exito(fila, hoja):
    hoja.cell(row=fila, column=3, value="✓")

def validar_numero(numero):
    numero = str(numero).strip().replace(" ", "").replace("-", "")
    if not numero:
        return False
    if len(numero) < 8:
        return False
    return numero if numero.startswith("+") else f"+{numero}"

def cargar_mensajes(wb):
    if "Mensajes" not in wb.sheetnames:
        raise ValueError("No existe la hoja 'Mensajes'")
    
    hoja_mensajes = wb["Mensajes"]
    mensaje_base = hoja_mensajes['B1'].value
    mensaje_sin_nombre = hoja_mensajes['B2'].value
    
    if not mensaje_base or not mensaje_sin_nombre:
        raise ValueError("Celdas B1 o B2 vacías en hoja 'Mensajes'")
    
    return mensaje_base, mensaje_sin_nombre

def enviar_mensaje(numero, nombre, mensaje_base, mensaje_sin_nombre):
    try:
        driver.get(f"https://web.whatsapp.com/send?phone={numero}")
        
        input_box = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//div[@role="textbox"][@data-tab="10"]'))
        )
        
        if nombre and nombre.strip():
            mensaje = mensaje_base.format(nombre=nombre.strip())
        else:
            mensaje = mensaje_sin_nombre
        
        lineas = mensaje.split('\n')
        
        for linea in lineas:
            input_box.send_keys(linea)
            input_box.send_keys(Keys.SHIFT + Keys.ENTER)
        input_box.send_keys(Keys.ENTER)
        
        try:
            WebDriverWait(driver, 12).until(
                EC.presence_of_element_located((By.XPATH, '//span[@data-testid="msg-dblcheck" or @data-testid="msg-time"]'))
            )
            print(f"✓✓ Envío confirmado a {numero}")
            return True
        except:
            if lineas[0] in input_box.text:
                print(f"✗ Error de envío a {numero}")
                return False
            print(f"✓ Envío probablemente exitoso a {numero}")
            return True
            
    except Exception as e:
        print(f"✗ Error crítico con {numero}: {str(e)}")
        return False

try:
    wb = openpyxl.load_workbook(ARCHIVO_CONFIG)
    
    # Verificar existencia de hoja Matriz
    if "Matriz" not in wb.sheetnames:
        raise ValueError("No existe la hoja 'Matriz' para los contactos")
    
    min_pausa, max_pausa, pausa_larga = cargar_tiempos_espera(wb)
    print(f"⚙️ Pausas configuradas: Individual={min_pausa}-{max_pausa}s | Cada 50 mensajes={pausa_larga}s")
    
    mensaje_base, mensaje_sin_nombre = cargar_mensajes(wb)
    
    driver.get("https://web.whatsapp.com")
    WebDriverWait(driver, 180).until(
        EC.presence_of_element_located((By.XPATH, '//div[@role="textbox"]'))
    )
    
    hoja_contactos = wb["Matriz"]  # Cambiado a hoja específica "Matriz"
    
    if hoja_contactos.cell(row=1, column=3).value != "Estado":
        hoja_contactos.cell(row=1, column=3, value="Estado")
    
    ultima_fila = encontrar_ultima_fila(hoja_contactos)
    print(f"📖 Contactos a procesar: {ultima_fila - 1}")
    
    contador = 0
    
    for fila_idx in range(2, ultima_fila + 1):
        numero_crudo = hoja_contactos.cell(row=fila_idx, column=1).value
        nombre = str(hoja_contactos.cell(row=fila_idx, column=2).value).strip() if hoja_contactos.cell(row=fila_idx, column=2).value else ""
        
        numero_validado = validar_numero(numero_crudo)
        if not numero_validado:
            marcar_error(fila_idx, hoja_contactos)
            print(f"✗ Número inválido en fila {fila_idx}")
            wb.save(ARCHIVO_CONFIG)
            continue
            
        if enviar_mensaje(numero_validado, nombre, mensaje_base, mensaje_sin_nombre):
            marcar_exito(fila_idx, hoja_contactos)
        else:
            marcar_error(fila_idx, hoja_contactos)
        
        wb.save(ARCHIVO_CONFIG)
        contador += 1
        
        if contador % 50 == 0:
            print(f"⏳ Pausa larga de {pausa_larga}s...")
            time.sleep(pausa_larga)
        else:
            tiempo_pausa = random.uniform(min_pausa, max_pausa)
            time.sleep(tiempo_pausa)

except Exception as e:
    print(f"⚠ Error: {str(e)}")
    
finally:
    try:
        wb.save(ARCHIVO_CONFIG)
        wb.close()
    except Exception as e:
        print(f"⚠ Error al guardar: {str(e)}")
    driver.quit()
    print("✅ Proceso completado")